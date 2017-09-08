# FILE: 	NERD.py
#
# PROGRAM:	NetApp Environment Review Document (NERD)
# 
# AUTHOR: Kellen Bryan
#  
# SUMMARY: NERD class for NERD_modeler.py
#
# Copyright (c) 2017 Network Appliance, Inc.
# All rights reserved.

########## MODULE IMPORT ############################################## 

#alphabet
import argparse
import codecs
from datetime import datetime,tzinfo,date
from dateutil.relativedelta import relativedelta
import fileinput
import math
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Fill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import os.path
import re
import requests
import statistics
import sys

class NERD():

	serial_numbers = []

	def __init__(self, serial_numbers_list):
		"""initialize Environment by adding serial #'s' to list. Serial #'s are accepted as argument or file"""
		for num in serial_numbers_list:
			self.serial_numbers.append(num)

	def _aggr_capacity(self, asup_url_output):
		"""Returns aggregate capacity"""
		line 			= asup_url_output
		match 			= re.findall("aggr_allocated_kb>(.*?)</aggr_allocated_kb>", line, re.DOTALL)
		match_name 		= re.findall("aggr_name>(.*?)</aggr_name>", line, re.DOTALL)
		match_return 	= {} #dictionary -> Key: aggr name; Value: aggr capacity

		if match:
			length = len(match)
			
			for i in range(0, length):
				capacity_TB = (float(match[i]) / (pow(1024, 3)))
				capacity_TB_rounded = round(capacity_TB, 2)
				match_return[str(match_name[i])] = capacity_TB_rounded
			
			return match_return

	def _aggr_name(self, asup_url_output):
		"""Returns aggregate names"""
		line 	= asup_url_output
		match 	= re.findall("(?<=Aggregate )(.*?)(?= \()", line)

		if match:
			match_list = []

			for name in match:
				if name not in match_list:
					match_list.append(str(name))
				elif name in match_list:
					break

			return match_list

	def _aggr_util(self, asup_url_output):
		"""Returns percent of aggr utilized"""
		line 			= asup_url_output
		match 			= re.findall("aggr_used_pct>(.*?)</aggr_used_pct>", line, re.DOTALL)
		match_name 		= re.findall("aggr_name>(.*?)</aggr_name>", line, re.DOTALL)
		match_return 	= {} #dictionary -> Key: aggr name; Value: aggr percent util

		if match:
			length = len(match)
			
			for i in range(0, length):
				match_return[str(match_name[i])] = int(match[i])
			
			return match_return

	def _asup_id(self, asup_url_output):
		"""Returns asup id for rest API use"""
		line 	= asup_url_output
		match 	= re.search("asup_id>(.*?)</asup_id", line, re.DOTALL)

		if match:
			return str(match.group(1))
	
	def _asup_received_date(self, asup_url_output):
		"""Return date of last recieved ASUP data"""
		line 	= asup_url_output
		match 	= re.search("asup_received_date>(.*?)</asup_received_date>", line, re.DOTALL)

		if match:
			return str(match.group(1))

	def _biz_key(self, asup_url_output):
		"""Returns biz_key for rest api use"""
		line 	= asup_url_output
		match 	= re.search("biz_key>(.*?)</biz_key>", line, re.DOTALL)

		if match:
			return str(match.group(1))

	def _capacity_forecast(self, asup_url_output):
		"""Returns growth rate of each aggregate (TB/month)"""
		line 		= asup_url_output
		match 		= re.findall("CDATA\[(.*?)\]\]>", line, re.DOTALL)
		match_name 	= re.search("CDATA\[(.*?)\]\]>", line, re.DOTALL)

		if match:

			match_return 			= {}
			capacity_dictionary 	= {}
			current_capacity 		= {}
			ninty_percent_capacity 	= {}
			match_name_buffer 		= []
			match_name_buffer 		= re.findall("(\S+)", str(match_name.group(1)))

			#filling dictionary with aggr names as keys
			for i in range(5, len(match_name_buffer),10):
				match_return[match_name_buffer[i]] 			= 0
				capacity_dictionary[match_name_buffer[i]] 	= []

			aggr_count = len(match_return)

			data_group_flag = 0
			#fill lists with capacity values from last 24 weeks
			for data_group in match:
				data_group_buffer = []
				data_group_buffer = re.findall("(\S+)", str(data_group))

				for i in range(0,aggr_count):
					if (7+(i*10)) <= len(data_group_buffer) and data_group_buffer[5+(i*10)] in match_name_buffer:
						if data_group_flag == 0:
							ninty_percent_capacity[data_group_buffer[5+(i*10)]] = (float(data_group_buffer[6+(i*10)]) * 0.9)/(pow(1024,3))
							current_capacity[data_group_buffer[5+(i*10)]] = (float(data_group_buffer[7+(i*10)])/(pow(1024,3)))
						capacity_dictionary[data_group_buffer[5+(i*10)]].append(float(data_group_buffer[7+(i*10)]))
				data_group_flag = 1

			#Calculate growth rate (AAGR)
			for name in capacity_dictionary:
				growth_rate_list = []

				for i in range(len(capacity_dictionary[name])-1):
					growth_rate_list.append(((capacity_dictionary[name][i]-capacity_dictionary[name][i+1])/capacity_dictionary[name][i+1])*100)
				
				average_growth_rate = np.average(growth_rate_list)
				over_ninty = ninty_percent_capacity[name] - current_capacity[name]

				if average_growth_rate != 0:
					capacity_forecast = (ninty_percent_capacity[name]-current_capacity[name])/average_growth_rate
				else: 
					match_return[name] = "More than one year"
					continue
				
				if over_ninty <= 0:
					match_return[name] = "Already > 90"
				elif capacity_forecast < 0:
					match_return[name] = "On decreasing trend"
				elif capacity_forecast <= 12:
					match_return[name] = "This year"
				elif capacity_forecast <= 3:
					match_return[name] = "This quarter"
				elif capacity_forecast <= 2:
					match_return[name] = "Next month"
				elif capacity_forecast > 12:
					match_return[name] = "More than one year"

			return match_return

	def _cluster_name(self, asup_url_output):
		"""Returns client-defined cluster name"""
		line 	= asup_url_output
		match 	= re.search("cluster_name>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _disk_count(self, asup_url_output):

		line 		= asup_url_output
		match 		= re.search("<data>(.*?)</data>", line, re.DOTALL)
		match_name 	= re.findall("(?<=Aggregate )(.*?)(?= \()", line)

		if match:

			match_buffer 	= []
			match_buffer 	= re.findall("(\S+)", str(match.group(1)))
			match_return 	= {} #dictionary -> Key: aggr name; Value: plex names

			disk_counter	= 0
			raid_group_flag = 0
			name_value 		= -1
			name_list 		= []

 			#add unique aggr name keys and create name list
			for agg_name in match_name:
				if agg_name not in name_list:
					name_list.append(str(agg_name))
					match_return[str(agg_name)] = 0
				elif agg_name in match_buffer:
					break
			
			#find disk count for each aggr
			for word in match_buffer:
				if "parity" in word and raid_group_flag == 1 and name_value < len(name_list):
					match_return[name_list[name_value]] = str(disk_counter) + " disks"
					raid_group_flag = 2
					disk_counter = 0

				elif "data" in  word and raid_group_flag != 2:
					raid_group_flag = 1
					disk_counter += 1

				elif "Aggregate" in word:
					if raid_group_flag == 1:
						raid_group_flag = 0
						match_return[name_list[name_value]] = str(disk_counter) + " disks"
						disk_counter = 0
						name_value += 1
						
					else:
						name_value += 1
						raid_group_flag = 0
					

				elif "spare" in word and raid_group_flag != 2:
					match_return[name_list[name_value]] = str(disk_counter) + " disks"
					break
			
			return match_return

	def _disk_type_count(self, asup_url_output):
		"""Returns number of each disk type"""
		line 		= asup_url_output
		match 		= re.search("<data>(.*?)</data>", line, re.DOTALL)
		match_name 	= re.findall("(?<=Aggregate )(.*?)(?= \()", line)

		if match:

			match_buffer 	= []
			match_buffer 	= re.findall("(\S+)", str(match.group(1)))
			match_return 	= {} #dictionary -> Key: aggr name; Value: RAID Group count

			name_value 		= 0
			name_list 		= []
			ssd_counter 	= 0
			sas_counter 	= 0
			new_aggr_flag 	= 0
			start_flag 		= 0
			i 				= 0

 			#add unique aggr name keys and create name list
			for agg_name in match_name:
				if agg_name not in name_list:
					name_list.append(str(agg_name))
					match_return[str(agg_name)] = 0
				elif agg_name in match_buffer:
					break

			#Count number of each type of RAID layout
			for word in match_buffer:

				i+=1

				if "Aggregate" in word or "spare" in word:
					if start_flag != 0:
						new_aggr_flag = 0
						match_return[name_list[name_value]] = "({} SAS; {} SSD)".format(sas_counter, ssd_counter)
						name_value += 1
						if "spare" in word:
							break
					else:
						start_flag = 1

				elif "Type" in word:
					if new_aggr_flag == 0:
						sas_counter = 0
						ssd_counter = 0
						new_aggr_flag = 1

					raid_type = match_buffer[i + 32]
					if "SSD" in raid_type:
						ssd_counter += 1
					elif "SAS" in raid_type:
						sas_counter += 1

			return match_return

	def _fiscal_end(self):
		"""Returns a datetime object of the next fiscal year end (assumes it's the last Friday of the following April)"""
		today 		= datetime.now()
		month 		= "04"
		day 		= 30
		year	 	= today.year
		fiscal_end 	= None

		if today.month > 4:
			year += 1

		while not fiscal_end:
			fiscal_end_datetime = datetime.strptime(str(year) + "-" + str(month) + "-" + str(day),"%Y-%m-%d")
			if fiscal_end_datetime.weekday() == 4:
				fiscal_end = fiscal_end_datetime
			else:
				day -= 1

		return fiscal_end

	def _growth_rate_monthly(self, asup_url_output):
		"""Returns growth rate of each aggregate (%/month)"""
		line 		= asup_url_output
		match 		= re.findall("CDATA\[(.*?)\]\]>", line, re.DOTALL)
		match_name 	= re.search("CDATA\[(.*?)\]\]>", line, re.DOTALL)

		if match:

			match_return 		= {}
			capacity_dictionary = {}
			match_name_buffer 	= []
			match_name_buffer 	= re.findall("(\S+)", str(match_name.group(1)))

			#filling dictionary with aggr names as keys
			for i in range(5, len(match_name_buffer),10):
				match_return[match_name_buffer[i]] 			= 0
				capacity_dictionary[match_name_buffer[i]] 	= []

			aggr_count = len(match_return)

			#fill lists with capacity values from last x weeks
			for data_group in match:
				data_group_buffer = []
				data_group_buffer = re.findall("(\S+)", str(data_group))

				for i in range(0,aggr_count):
					if (7+(i*10)) <= len(data_group_buffer) and data_group_buffer[5+(i*10)] in match_name_buffer:
						capacity_dictionary[data_group_buffer[5+(i*10)]].append(float(data_group_buffer[7+(i*10)]))
			
			for name in capacity_dictionary:
				n = len(capacity_dictionary[name])

				#Calculate growth rate (CAGR)
				x = capacity_dictionary[name][0]/(capacity_dictionary[name][n-1])
				cagr_growth_rate = round((pow(x,(1/float(n)))-1)*100,2)
				
				#Calculate growth rate (AAGR)
				growth_rate_list = []
				for i in range(n-1):
					growth_rate_list.append(((capacity_dictionary[name][i]-capacity_dictionary[name][i+1]) /capacity_dictionary[name][i+1])*100)
				
				average_growth_rate = round(np.average(growth_rate_list), 2) #growth rate per week
				if average_growth_rate >= 100:
					match_return[name] = "Over 100%. Check ASUP for details."
				else:
					match_return[name] = (average_growth_rate*4) #growth rate per month

			return match_return

	def _growth_tb_monthly(self, asup_url_output):
		"""Returns growth rate of each aggregate (TB/month)"""
		line 		= asup_url_output
		match 		= re.findall("CDATA\[(.*?)\]\]>", line, re.DOTALL)
		match_name 	= re.search("CDATA\[(.*?)\]\]>", line, re.DOTALL)

		if match:

			match_return 		= {}
			capacity_dictionary = {}
			match_name_buffer 	= []
			match_name_buffer 	= re.findall("(\S+)", str(match_name.group(1)))

			#filling dictionary with aggr names as keys
			for i in range(5, len(match_name_buffer),10):
				match_return[match_name_buffer[i]] 			= 0
				capacity_dictionary[match_name_buffer[i]] 	= []

			aggr_count = len(match_return)

			#fill lists with capacity values from last 24 weeks
			for data_group in match:
				data_group_buffer = []
				data_group_buffer = re.findall("(\S+)", str(data_group))

				for i in range(0, aggr_count):
					if (7+(i*10)) <= len(data_group_buffer) and data_group_buffer[5+(i*10)] in match_name_buffer:
						capacity_dictionary[data_group_buffer[5+(i*10)]].append(float(data_group_buffer[7+(i*10)]))

			#calculate average growth in TB per month
			growth_tb = []
			for name in capacity_dictionary:
				for i in range(len(capacity_dictionary[name])-1):
					growth_tb.append(capacity_dictionary[name][i]-capacity_dictionary[name][i+1])
				average_difference = round((np.average(growth_tb)/(pow(1024, 3))), 2) #average growth per week
				match_return[name] = (average_difference * 4) #average growth per month
			return match_return

	def _host_name(self, asup_url_output):
		"""Returns product host name"""
		line 	= asup_url_output
		match 	= re.search("hostname>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _location(self, asup_url_output):
		"""Returns customer location. Location defines different pages"""
		line 	= asup_url_output
		match 	= re.search("site_name>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _performance_iops(self, asup_url_output):
		"""Returns iops and cpu busy % (std_dev + avg) for previous week"""
		line 			= asup_url_output
		fcp_ops_match 	= re.findall("fcp_ops.*?<counterValue>(.*?)</counterValue>", line, re.DOTALL)
		iscsi_ops_match = re.findall("iscsi_ops.*?<counterValue>(.*?)</counterValue>", line, re.DOTALL)
		cifs_ops_match 	= re.findall("cifs_ops.*?<counterValue>(.*?)</counterValue>", line, re.DOTALL)
		nfs_ops_match 	= re.findall("nfs_ops.*?<counterValue>(.*?)</counterValue>", line, re.DOTALL)
		cpu_busy_match 	= re.findall("cpu_busy.*?<counterValue>(.*?)</counterValue>", line, re.DOTALL)

		if fcp_ops_match or iscsi_ops_match or cifs_ops_match or nfs_ops_match or cpu_busy_match:

			fcp_ops 	= []
			iscsi_ops 	= []
			cifs_ops 	= []
			nfs_ops 	= []
			cpu_busy 	= []

			length = len(fcp_ops_match)

			for i in range(0, length):
				fcp_ops.append(float(fcp_ops_match[i]))
				iscsi_ops.append(float(iscsi_ops_match[i]))		
				cifs_ops.append(float(cifs_ops_match[i]))		
				nfs_ops.append(float(nfs_ops_match[i]))		
				cpu_busy.append(float(cpu_busy_match[i]))		

			#calculate averages and std dev over last week
			fcp_avg 		= sum(fcp_ops)/length
			fcp_std_dev 	= statistics.stdev(fcp_ops)

			iscsi_avg 		= sum(iscsi_ops)/length
			iscsi_std_dev 	= statistics.stdev(iscsi_ops)

			cifs_avg 		= sum(cifs_ops)/length
			cifs_std_dev 	= statistics.stdev(cifs_ops)

			nfs_avg		 	= sum(nfs_ops)/length
			nfs_std_dev 	= statistics.stdev(nfs_ops)

			cpu_avg 		= sum(cpu_busy)/length
			cpu_std_dev 	= statistics.stdev(cpu_busy)

			return_list = []
			return_list.append(round( (cpu_avg+cpu_std_dev), 2))
			return_list.append(round( (cifs_avg+cifs_std_dev), 2))
			return_list.append(round( (fcp_avg+fcp_std_dev), 2))
			return_list.append(round( (iscsi_avg+iscsi_std_dev), 2))
			return_list.append(round( (nfs_avg+nfs_std_dev), 2))
			
			return return_list

		no_match_list = ['No Data Available', 'No Data Available', 'No Data Available', 'No Data Available', 'No Data Available']
		return no_match_list

	def _raid_group_count(self, asup_url_output):
		"""Returns RAID group set-up"""
		line 		= asup_url_output
		match_name 	= re.findall("(?<=Aggregate )(.*?)(?= \()", line)
		match_raid	= re.findall("(?<=RAID group)(.*?)(?=\()", line)

		if match_name and match_raid:

			match_buffer = {} #dictionary -> Key: aggr name; Value: raid group names
			for agg_name in match_name:
				if agg_name not in match_buffer: #add unique aggr name keys
					match_buffer[str(agg_name)] = [] 
					for raid in match_raid:
						if agg_name in raid and raid not in match_buffer[str(agg_name)]: #add plex names to corresponding aggr names
							match_buffer[str(agg_name)].append(str(raid))
				elif agg_name in match_buffer:
					break
			
			return match_buffer
	
	def _raid_type(self, asup_url_output):
		"""Returns percent of aggr utilized"""
		line 		= asup_url_output
		match 		= re.findall("aggr_raid_type>(.*?)</aggr_raid_type>", line, re.DOTALL)
		match_name 	= re.findall("aggr_name>(.*?)</aggr_name>", line, re.DOTALL)
		
		if match:

			match_return = {} #dictionary -> Key: aggr name; Value: raid type
			length = len(match)
			
			for i in range(0, length):
				match_return[str(match_name[i])] = str(match[i])
			
			return match_return

	def _system_id(self, asup_url_output):
		"""Returns system id number"""
		line 	= asup_url_output
		match 	= re.search("system_id>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _system_model(self, asup_url_output):
		"""Returns system model/controller"""
		line 	= asup_url_output
		match 	= re.search("sys_model>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _serial_number(self, asup_url_output):
		"""Returns product serial number"""
		line 	= asup_url_output
		match 	= re.search("sys_serial_no>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _system_version(self, asup_url_output):
		"""Returns OS version"""
		line 	= asup_url_output
		match 	= re.search("sys_version>(.*?)<", line)

		if match:

			return str(match.group(1))

	def _volume_iops(self, asup_url_output):
		"""Returns volume IOPs (mean)"""
		line 			= asup_url_output
		name_match 		= re.findall("instance_name>(.*?)<", line)
		counter_match 	= re.findall("counter_stat_value>(.*?)<", line)

		if counter_match:

			volume_iops_dic 	= {}
			match_name_buffer 	= []
			iops_count_buffer 	= []

			for name in name_match:
				if " " not in name:
					match_name_buffer.append(str(name))
			for iops in counter_match:
				if " " not in iops:
					iops_count_buffer.append(str(iops))

			for i in range(0, len(match_name_buffer)-1, 2):
				volume_iops_dic[str(match_name_buffer[i])] = abs(round(float(iops_count_buffer[i]), 2))

			return volume_iops_dic 

	def _warranty_status(self, asup_url_output):
		"""Returns warranty expiration date"""
		line 	= asup_url_output
		match 	= re.search("warranty_end_date>(.*?)</warranty_end_date>", line)

		if match:

			string_date = str(match.group(1)).replace("-","")
			date 		= datetime.strptime(string_date, '%Y%m%d').date()
			return date

	
		
	

	

	

#ASUP APIS
#http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/biz_key/C%7C93E0D750-5BDF-11E4-9F60-123478563412%7C8499809755%7C721545000241/object_list
#http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/biz_key/C%7C93E0D750-5BDF-11E4-9F60-123478563412%7C8499809755%7C721545000241/list