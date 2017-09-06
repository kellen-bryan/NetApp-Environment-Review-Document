# FILE: 	NERD_Modeler.py
#
# PROGRAM:	NetApp Environment Review Document (NERD)
#
# AUTHOR: 	Kellen Bryan
#
# SUMMARY: 	NERD mines relevant information from ASUP about client storage environments
# 			and organizes the data in a user-friendly worksheet.
#
# USAGE:	NERD provides the NetApp Sales Team with organized metrics to proactively 
# 			make recommendations on improvements to better serve the customer and also 
# 			faciliates education clients on the overall performance of their storage environment.
#
# Copyright (c) 2017 Network Appliance, Inc.
# All rights reserved.  


#################### MODULE IMPORT #################### 

import argparse
import codecs
from datetime import date, datetime, tzinfo, timedelta
from dateutil.relativedelta import relativedelta
import fileinput
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo
from operator import itemgetter
import os.path
import re
import requests
import sys
import time

#NERD class file
from NERD import *

t0 = time.time()



#################### SHEET HEADERS #################### 

def _location_sheet_column_headers():
	location_sheet_column_headers = []
	
	location_sheet_column_headers.append("Location")
	location_sheet_column_headers.append("Cluster Name")
	location_sheet_column_headers.append("Host Name")
	location_sheet_column_headers.append("Controller")
	location_sheet_column_headers.append("Serial Number")
	location_sheet_column_headers.append("OS Version")
	location_sheet_column_headers.append("Application")
	location_sheet_column_headers.append("Warranty End Date")

	return location_sheet_column_headers

def _raid_info_sheet_column_headers():
	raid_info_sheet_column_headers = []
	
	raid_info_sheet_column_headers.append("Cluster Name")
	raid_info_sheet_column_headers.append("Host Name")
	raid_info_sheet_column_headers.append("Controller")
	raid_info_sheet_column_headers.append("Serial Number")
	raid_info_sheet_column_headers.append("OS Version")
	raid_info_sheet_column_headers.append("Aggr Name")
	raid_info_sheet_column_headers.append("RAID Type")
	raid_info_sheet_column_headers.append("RAID Layout")
	raid_info_sheet_column_headers.append("Data Disks Per RAID Group")
	raid_info_sheet_column_headers.append("Aggr Capacity (TB)")
	raid_info_sheet_column_headers.append("Aggr Util (%)")

	return raid_info_sheet_column_headers

def _capacity_trending_sheet_column_headers():
	capacity_trending_sheet_column_headers = []

	capacity_trending_sheet_column_headers.append("Cluster Name")
	capacity_trending_sheet_column_headers.append("Host Name")
	capacity_trending_sheet_column_headers.append("Controller")
	capacity_trending_sheet_column_headers.append("Serial Number")
	capacity_trending_sheet_column_headers.append("Aggr Name")
	capacity_trending_sheet_column_headers.append("Average Growth (TB/month)")
	capacity_trending_sheet_column_headers.append("Average Growth Rate (%/month)")
	capacity_trending_sheet_column_headers.append("90% Capacity")

	return capacity_trending_sheet_column_headers

def _performance_sheet_column_headers_1():
	performance_sheet_column_headers_1 = []

	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("st_dev + avg")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("")
	performance_sheet_column_headers_1.append("Max")

	return performance_sheet_column_headers_1

def _performance_sheet_column_headers_2():
	performance_sheet_column_headers_2 = []

	performance_sheet_column_headers_2.append("Cluster Name")
	performance_sheet_column_headers_2.append("Host Name")
	performance_sheet_column_headers_2.append("Controller")
	performance_sheet_column_headers_2.append("Serial Number")
	performance_sheet_column_headers_2.append("CPU%")
	performance_sheet_column_headers_2.append("CIFS (IOPS)")
	performance_sheet_column_headers_2.append("FCP (IOPS)")
	performance_sheet_column_headers_2.append("iSCI (IOPS)")
	performance_sheet_column_headers_2.append("NFS (IOPS)")
	performance_sheet_column_headers_2.append("Latency (ms)")

	return performance_sheet_column_headers_2

def _volumes_sheet_column_headers():
	volumes_sheet_column_headers = []

	volumes_sheet_column_headers.append("Cluster Name")
	volumes_sheet_column_headers.append("Host Name")
	volumes_sheet_column_headers.append("Controller")
	volumes_sheet_column_headers.append("Serial Number")
	volumes_sheet_column_headers.append("Volume Name")
	volumes_sheet_column_headers.append("IOPS (avg)")

	return volumes_sheet_column_headers



########################################################## MAIN ########################################################## 

parser = argparse.ArgumentParser(description='Serial Number Input')
parser.add_argument('-serial-numbers', help='Enter serial numbers seperated by commas (no spaces)')
parser.add_argument('-file-path', help='Enter file path')
args = parser.parse_args()

serial_numbers_list = [] 



#################### ARGUMENT HANDLING ####################

#Given file
if args.file_path:
	#check to make sure file path exists
	if not os.path.exists(args.file_path):
		parser.error("The file %s does not exist" % args.file_path)
		sys.exit(1)

	#Given file
	print "Reading from file.."
	with open(args.file_path, 'r') as f:
		for line in f:
			serial_num = line.strip()
			if serial_num not in serial_numbers_list and serial_num != '':
				serial_numbers_list.append(serial_num)
	for num in serial_numbers_list:
		print num

#Manually enter serial numbers on Command Line
elif args.serial_numbers:
	print "Serial numbers from STDIN..."
	serial_numbers_list = args.serial_numbers.split(',')
	for num in serial_numbers_list:
		print num

#Manual entry after starting program
else:
	print "Enter list of serial numbers seperated by commas: "
	string_input = raw_input()
	serial_numbers_list = string_input.split(',')
	serial_numbers_list = [int(num) for num in serial_numbers_list]
	for num in serial_numbers_list:
		print num



#################### CREATE EXCEL DOC ####################

wb 				= Workbook()
dimensions 		= {}



#################### RETRIEVE DATA FROM ASUP #################### 

#Create dictionary for each page
location_dictionary 			= {}
cluster_dictionary 				= {}
capacity_trending_dictionary 	= {}
performance_dictionary 			= {}
volumes_dictionary 				= {}

no_cluster_count = 0
today = date.today()
start_date = today - timedelta(weeks=12)
print "START DATE: "
print start_date

#Get data from REST APIs
for serial_number in serial_numbers_list:

	print "GETTING INFO FOR SERIAL NUM: " + str(serial_number)

	#API url with storage environment overview
	asup_overview_url 			= "http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/sys_serial_no/" + str(serial_number)
	asup_overview_url_output 	= requests.get(asup_overview_url).text
	current_page 				= NERD(asup_overview_url_output)

	asup_id 			= current_page._asup_id(asup_overview_url_output)
	asup_received_date 	= current_page._asup_received_date(asup_overview_url_output)
	biz_key 			= current_page._biz_key(asup_overview_url_output)
	cluster_name 		= current_page._cluster_name(asup_overview_url_output)
	host_name 			= current_page._host_name(asup_overview_url_output)
	location 			= current_page._location(asup_overview_url_output)
	os_version 			= current_page._system_version(asup_overview_url_output)
	system_id			= current_page._system_id(asup_overview_url_output)
	system_model		= current_page._system_model(asup_overview_url_output)
	serial_number 		= current_page._serial_number(asup_overview_url_output)
	warranty_status 	= current_page._warranty_status(asup_overview_url_output)

	#API url with storage environment configuration info (SYSCONFIG-R)
	asup_sysconfigR_url 		= "http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/sys_serial_no/" + str(serial_number) + "/section_view/SYSCONFIG-R"
	asup_sysconfigR_url_output 	= requests.get(asup_sysconfigR_url).text
	current_page 				= NERD(asup_sysconfigR_url_output)
	
	aggr_name 			= current_page._aggr_name(asup_sysconfigR_url_output)
	raid_group_count 	= current_page._raid_group_count(asup_sysconfigR_url_output)
	disk_count 			= current_page._disk_count(asup_sysconfigR_url_output)
	disk_type_count 	= current_page._disk_type_count(asup_sysconfigR_url_output)

	#API url to get system IOP info
	asup_iops_url 			= "http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/sys_serial_no/" + str(serial_number) + "/object/system/counter_name/cifs_ops,nfs_ops,fcp_ops,iscsi_ops,cpu_busy/cvc"
	asup_iops_url_output 	= requests.get(asup_iops_url).text
	current_page 			= NERD(asup_iops_url_output)

	performance_iops 		= current_page._performance_iops(asup_iops_url_output)

	#API url to get volume IOPS
	asup_volume_iops_url = "http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/system_id/" + str(system_id) + "/sys_serial_no/" + str(serial_number) + "/start_date/" + str(start_date) + "/end_date/" + str(today) + "/object/volume/counter_name/total_ops/stat/mean/csc/"
	asup_volume_iops_url_output = requests.get(asup_volume_iops_url).text
	current_page = NERD(asup_volume_iops_url_output)

	volume_iops = current_page._volume_iops(asup_volume_iops_url_output)

	#API url with growth rate info for past 24 weeks (DF-A)
	asup_DFA_url 		= "http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/sys_serial_no/" + str(serial_number) + "/start_date/" + str(start_date) + "/end_date/" + str(today) + "/section_view/DF-A"
	asup_DFA_url_output = requests.get(asup_DFA_url).text
	current_page 		= NERD(asup_DFA_url_output)

	growth_tb_monthly 	= current_page._growth_tb_monthly(asup_DFA_url_output)
	growth_rate_monthly = current_page._growth_rate_monthly(asup_DFA_url_output)
	capacity_forecast 	= current_page._capacity_forecast(asup_DFA_url_output)

	#API url with aggregate and raid info 
	asup_aggregate_info_url 		= "http://restprd.corp.netapp.com/asup-rest-interface/ASUP_DATA/client_id/test/asup_id/" + str(asup_id) + "/object_view/AGGREGATE"
	asup_aggregate_info_url_output 	= requests.get(asup_aggregate_info_url).text
	current_page 					= NERD(asup_aggregate_info_url_output)

	#Try until webpage successfully reached
	request_flag = 0
	try_count = 0
	while request_flag != 1:
		error_match = re.search("(Error)", asup_aggregate_info_url_output)
		if error_match:
			if try_count > 10:
				print "Server not responding for serial number: " + str(serial_number)
				print "Exiting program...Please try again"
				sys.exit()
			try_count += 1
			asup_aggregate_info_url_output = requests.get(asup_aggregate_info_url).text
		else:
			request_flag = 1
	
	aggr_capacity = current_page._aggr_capacity(asup_aggregate_info_url_output)
	aggr_util = current_page._aggr_util(asup_aggregate_info_url_output)
	raid_type = current_page._raid_type(asup_aggregate_info_url_output)


	#Fill dictionaries
	if location not in location_dictionary:
		location_dictionary[location] = {}

	if cluster_name == None:
		cluster_name = "No_Cluster_Name_{0}".format(no_cluster_count)
		no_cluster_count+=1

	#if cluster_name not in location_dictionary:
	if cluster_name not in cluster_dictionary or "No_Cluster" in cluster_name:
		cluster_dictionary[cluster_name] 			= {}
		capacity_trending_dictionary[cluster_name] 	= {}
		performance_dictionary[cluster_name] 		= {}
		location_dictionary[location][cluster_name] = []
		volumes_dictionary[cluster_name] 			= {}

	if host_name not in cluster_dictionary[cluster_name]:
		cluster_dictionary[cluster_name][host_name] 			= []
		capacity_trending_dictionary[cluster_name][host_name] 	= []
		performance_dictionary[cluster_name][host_name] 		= []
		volumes_dictionary[cluster_name][host_name] 			= [] 
	
	for name in aggr_name:
		growth_tb = growth_tb_monthly[name]
		growth_rate = growth_rate_monthly[name]
		capacity = capacity_forecast[name]
		disks = disk_count[name]
		aggr_cap = aggr_capacity[name]
		raid_layout = str(len(raid_group_count[name])) + " groups" + " " + str(disk_type_count[name])
		aggr_util_percent = aggr_util[name]
		type_raid = raid_type[name]

		cluster_dictionary[cluster_name][host_name].append([system_model, serial_number, os_version, name, type_raid, raid_layout, disks, aggr_cap, aggr_util_percent])
		capacity_trending_dictionary[cluster_name][host_name].append([system_model, serial_number, name, growth_tb, growth_rate, capacity])

	for name in volume_iops:
		iops = volume_iops[name]

		volumes_dictionary[cluster_name][host_name].append([system_model, serial_number, name, iops])

	performance_dictionary[cluster_name][host_name].append(system_model)
	performance_dictionary[cluster_name][host_name].append(serial_number)
	performance_dictionary[cluster_name][host_name].extend(performance_iops)
	location_dictionary[location][cluster_name].append([host_name, system_model, serial_number, os_version, "", warranty_status])



#################### DOCUMENT OVERVIEW SHEET #################### 

overview_sheet = wb.active
overview_sheet.title = "Document Overview"

overview_sheet.merge_cells('A1:B1')

document_overview_sheet_column_headers_1 = []
document_overview_sheet_column_headers_1.append("NetApp Environment Review Document (NERD)")
document_overview_sheet_column_headers_1.append("")
document_overview_sheet_column_headers_1.append("ASUP Received Date: {}".format(asup_received_date))

document_overview_sheet_column_headers_2 = []
document_overview_sheet_column_headers_2.append("Sheet")
document_overview_sheet_column_headers_2.append("Contents")
document_overview_sheet_column_headers_2.append("Usage")

location_info = []
location_info.append("Location")
location_info.append("General location information. Includes cluster names, host names, controller type, OS version, etc.")
location_info.append("Allows user to filter and examine general storage environment set-up at particular locations.")

raid_info = []
raid_info.append("RAID Info")
raid_info.append("Details of RAID set-up and aggregate utilization.")
raid_info.append("Sheet is organized by cluster name and gives details on RAID set-up on each aggregate, including RAID type, disk count, aggregate capacity, etc.")

capacity_trending_info = []
capacity_trending_info.append("Capacity Trending")
capacity_trending_info.append("Growth rate information for each cluster.")
capacity_trending_info.append("View growth rates to predict upsell opportunities for clients")

performance_info = []
performance_info.append("Performance")
performance_info.append("Overall performance of environment. Includes IOPS, max latency, etc.")
performance_info.append("View overall health of client storage system to proactively make recommendations on upgrades and improvements.")

overview_sheet.append(document_overview_sheet_column_headers_1)
overview_sheet.append(document_overview_sheet_column_headers_2)
overview_sheet.append(location_info)
overview_sheet.append(raid_info)
overview_sheet.append(capacity_trending_info)
overview_sheet.append(performance_info)

#create proper widths for cells
for row in overview_sheet.iter_rows(max_row=1):
	for cell in row:
		dimensions[cell.column] = 50
for col, value in dimensions.items():
	overview_sheet.column_dimensions[col].width = value

#Format header rows
for row in overview_sheet.iter_rows(min_row=1, max_row=2, max_col = 3):
	for cell in row:
		cell.fill 		= PatternFill(start_color="000080", fill_type="solid")
		cell.font 		= Font(bold=True, color="ffffff")
		cell.alignment 	= Alignment(horizontal='center')
		cell.border 	= Border(
			left		= Side(style 	= 'thin',
								color 	= "ffffff"),
			right 		= Side(style 	= 'thin',
								color 	= "ffffff"),
			top			= Side(style 	= 'thin',
								color 	= "ffffff"),
			bottom		= Side(style 	= 'thin',
								color 	= "ffffff"))

for row in overview_sheet.iter_rows(min_row=2):
	for cell in row:
		cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

for col in overview_sheet.iter_cols(min_row=3, max_col=1):
	for cell in col:
		cell.border = Border(left = Side(style = 'thick'))
for col in overview_sheet.iter_cols(min_col=4, max_col=4):
	for cell in col:
		cell.border = Border(left = Side(style = 'thick'))
for row in overview_sheet.iter_rows(min_row=7, max_row=7, max_col=3):
	for cell in row:
		cell.border = Border(top = Side(style = 'thick'))



#################### LOCATION SHEET #################### 

#Add column headers to sheet
sheet = wb.create_sheet("Locations")
location_sheet_column_headers = _location_sheet_column_headers()
sheet.append(location_sheet_column_headers)

#create proper widths for cells
for row in sheet.iter_rows(max_row = 1):
	for cell in row:
		if cell.value:
			dimensions[cell.column] = max((dimensions.get(cell.column, 0), len(str(cell.value))+12))
for col, value in dimensions.items():
	sheet.column_dimensions[col].width = value

#Alphabetize locations for proper sheet location
location_list = []
for location in location_dictionary:
	location_list.append(location)
location_list.sort()

color_flag = 0
row_flag = 0

#Populate sheet
for location in location_list:
	location_flag = 1
	location_row_count = 0

	#Alphabetize cluster
	cluster_list = []
	for cluster_name in location_dictionary[location]:
		if cluster_name != None:
			cluster_list.append(cluster_name)
	cluster_list.sort()

	
	#Alphabetize by Host Name
	for cluster in cluster_list:

		row_info_list = []
		for row_info in location_dictionary[location][cluster]:
			row_info_list.append(row_info)
		row_info_list.sort()

		#Populate spreadsheet
		all_info = []
		for item in row_info_list:
			all_info.append(location)
			all_info.append(cluster)
			for thing in item:
				all_info.append(thing)
			sheet.append(all_info)
			all_info = []
	 
			
		#alternate colors for each cluster (white/blue)
		for row in sheet.iter_rows(min_row=2, max_col=8):
			if row[1].value == cluster:
				for cell in row:
					if color_flag == 1:
						cell.fill = PatternFill(start_color="ffffff", fill_type="solid")
					if color_flag == 0:
						cell.fill = PatternFill(start_color="add8e6", fill_type="solid")
		if color_flag == 0:
			color_flag = 1
		elif color_flag == 1:
			color_flag = 0

	#get number of rows for each location
	for row in sheet.iter_rows(min_row=2, max_col=1):
		for cell in row:
			if cell.value == location:
				location_row_count += 1	
	
	#apply borders as per location 
	for row in sheet.iter_rows(min_row=2, max_col=8):
		if row[0].value == location and location_flag < location_row_count:	
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thin'))
				else:
					cell.border = Border(
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom 	= Side(style='thin'))

			location_flag += 1
				
		elif row[0].value == location and location_flag == location_row_count: 
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left 	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thick'))

				else:
					cell.border = Border(	
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom	= Side(style='thick'))

#freeze top row
sheet.freeze_panes='A2'

#add autofilter
column_letter 			= get_column_letter(len(location_sheet_column_headers))
highest_row 			= sheet.max_row
last_cell 				= column_letter + str(highest_row)
sheet.auto_filter.ref 	= "A1:" + last_cell


#format cells
for col in sheet.iter_cols(min_row=1, max_row=1, max_col = 8):
		for cell in col:
			cell.fill 		= PatternFill(start_color="000080", fill_type="solid")
			cell.font 		= Font(bold=True, color="ffffff")
			cell.alignment 	= Alignment('center')
			cell.border 	= Border(
				left		= Side(style ='thin'),
				right 		= Side(style ='thin'),
				top			= Side(style ='thin'),
				bottom		= Side(style ='thin'))

#check warranty end date
fiscal_year_end = current_page._fiscal_end()
print fiscal_year_end
for col in sheet.iter_cols(min_col=8 ,max_col=8,min_row=2):
	for cell in col:
		if cell.value <= fiscal_year_end:
			cell.fill = PatternFill(start_color="ffff00", fill_type="solid")
		elif cell.value <= datetime.now():
			cell.fill = PatternFill(start_color="ff0000", fill_type="solid")



#################### RAID SHEET #################### 

#Create sheet
sheet2 = wb.create_sheet("Raid Info")
raid_info_sheet_column_headers = _raid_info_sheet_column_headers()
sheet2.append(raid_info_sheet_column_headers)

#Alphabetize clusters
cluster_list = []
for cluster in cluster_dictionary:
	cluster_list.append(cluster)
cluster_list.sort()

color_flag = 0
row_flag = 0

#Populate worksheet 
for cluster in cluster_list:
	cluster_flag = 1
	cluster_row_count = 0

	#Alphabetize host names
	host_name_list = []
	for host_name in cluster_dictionary[cluster]:
		if host_name != None:
			host_name_list.append(host_name)
	host_name_list.sort()

	#Alphabetize by Host Name
	for host_name in host_name_list:

		row_info_list = []
		for row_info in cluster_dictionary[cluster][host_name]:
			row_info_list.append(row_info)
		row_info_list.sort()

		#Populate spreadsheet
		all_info = []
		for row in row_info_list:
			all_info.append(cluster)
			all_info.append(host_name)
			for item in row:
				all_info.append(item)
			sheet2.append(all_info)
			all_info = []
	 
		
		#alternate colors for each cluster (white/blue)
		for row in sheet2.iter_rows(min_row=2, max_col=11):
			if row[0].value == cluster:
				for cell in row:
					if color_flag == 1:
						cell.fill = PatternFill(start_color="ffffff", fill_type="solid")
					if color_flag == 0:
						cell.fill = PatternFill(start_color="add8e6", fill_type="solid")
	if color_flag == 0:
		color_flag = 1
	elif color_flag == 1:
		color_flag = 0

	#get number of rows for each cluster
	for row in sheet2.iter_rows(min_row=2, max_col=1):
		for cell in row:
			if cell.value == cluster:
				cluster_row_count += 1	

	#apply borders as per cluster 
	for row in sheet2.iter_rows(min_row=2, max_col=11):
		if row[0].value == cluster and cluster_flag < cluster_row_count:	
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thin'))
				else:
					cell.border = Border(
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom 	= Side(style='thin'))

			cluster_flag += 1
				
		elif row[0].value == cluster and cluster_flag == cluster_row_count: 
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left 	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thick'))

				else:
					cell.border = Border(	
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom	= Side(style='thick'))

#add thin borders to all cells
for row in sheet2.iter_rows(min_row=1, max_row=1, max_col = 11):
	for cell in row:
		cell.fill 		= PatternFill(start_color="000080", fill_type="solid")
		cell.font 		= Font(bold=True, color="ffffff")
		cell.alignment 	= Alignment('center')
		cell.border 	= Border(
			left		= Side(style ='thin'),
			right 		= Side(style ='thin'),
			top			= Side(style ='thin'),
			bottom		= Side(style ='thin'))

#create proper widths for cells
for row in sheet2.iter_rows(max_row = 1):
	for cell in row:
		if cell.value:
			dimensions[cell.column] = max((dimensions.get(cell.column, 0), len(str(cell.value))+12))
for col, value in dimensions.items():
	sheet2.column_dimensions[col].width = value

for col in sheet2.iter_cols(min_col = 11, min_row =2):
	for cell in col:
		if cell.value >= 90:
			cell.fill = PatternFill(start_color="ff0000", fill_type="solid")
		elif cell.value >= 80:
			cell.fill = PatternFill(start_color="ffff00", fill_type="solid")

#freeze top row
sheet2.freeze_panes='A2'

#add autofilter
column_letter 			= get_column_letter(len(raid_info_sheet_column_headers))
highest_row 			= sheet2.max_row
last_cell 				= column_letter + str(highest_row)
sheet2.auto_filter.ref 	= "A1:"  + last_cell



#################### CAPACITY TRENDING SHEET #################### 

#Create sheet
sheet3 = wb.create_sheet("Capacity Trending")
capacity_trending_sheet_column_headers = _capacity_trending_sheet_column_headers()
sheet3.append(capacity_trending_sheet_column_headers)

#Alphabetize clusters
cluster_list = []
for cluster in capacity_trending_dictionary:
	cluster_list.append(cluster)
cluster_list.sort()

color_flag = 0
row_flag = 0

#Populate worksheet 
for cluster in cluster_list:
	cluster_flag = 1
	cluster_row_count = 0

	#Alphabetize host names
	host_name_list = []
	for host_name in capacity_trending_dictionary[cluster]:
		if host_name != None:
			host_name_list.append(host_name)
	host_name_list.sort()
	
	#Alphabetize by Host Name
	for host_name in host_name_list:

		row_info_list = []
		for row_info in capacity_trending_dictionary[cluster][host_name]:
			row_info_list.append(row_info)
		row_info_list.sort()

		#Populate spreadsheet
		all_info = []
		for row in row_info_list:
			all_info.append(cluster)
			all_info.append(host_name)
			for item in row:
				all_info.append(item)
			sheet3.append(all_info)
			all_info = []
	 
		
		#alternate colors for each cluster (white/blue)
		for row in sheet3.iter_rows(min_row=2, max_col=8):
			if row[0].value == cluster:
				for cell in row:
					if color_flag == 1:
						cell.fill = PatternFill(start_color="ffffff", fill_type="solid")
					if color_flag == 0:
						cell.fill = PatternFill(start_color="add8e6", fill_type="solid")
	if color_flag == 0:
		color_flag = 1
	elif color_flag == 1:
		color_flag = 0

	#get number of rows for each cluster
	for row in sheet3.iter_rows(min_row=2, max_col=1):
		for cell in row:
			if cell.value == cluster:
				cluster_row_count += 1	

	#apply borders as per cluster 
	for row in sheet3.iter_rows(min_row=2, max_col=8):
		if row[0].value == cluster and cluster_flag < cluster_row_count:	
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thin'))
				else:
					cell.border = Border(
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom 	= Side(style='thin'))

			cluster_flag += 1
				
		elif row[0].value == cluster and cluster_flag == cluster_row_count: 
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left 	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thick'))

				else:
					cell.border = Border(	
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom	= Side(style='thick'))

#add thin borders to all cells
for row in sheet3.iter_rows(min_row=1, max_row=1, max_col = 8):
	for cell in row:
		cell.fill 		= PatternFill(start_color="000080", fill_type="solid")
		cell.font 		= Font(bold=True, color="ffffff")
		cell.alignment 	= Alignment('center')
		cell.border 	= Border(
			left		= Side(style ='thin'),
			right 		= Side(style ='thin'),
			top			= Side(style ='thin'),
			bottom		= Side(style ='thin'))

#create proper widths for cells
for row in sheet3.iter_rows(max_row = 1):
	for cell in row:
		if cell.value:
			dimensions[cell.column] = max((dimensions.get(cell.column, 0), len(str(cell.value))+12))
for col, value in dimensions.items():
	sheet3.column_dimensions[col].width = value

for col in sheet3.iter_cols(min_col=8, min_row=2):
	for cell in col:
		if cell.value == "Already > 90" or cell.value == "This year" or cell.value == "This quarter" or cell.value == "Next month":
			cell.font = Font(bold=True, color="ff0000")


#freeze top row
sheet3.freeze_panes='A2'

#add autofilter
column_letter 			= get_column_letter(len(capacity_trending_sheet_column_headers))
highest_row 			= sheet3.max_row
last_cell 				= column_letter + str(highest_row)
sheet3.auto_filter.ref 	= "A1:"  + last_cell



#################### PERFORMANCE SHEET #################### 

#Create sheet
sheet4 = wb.create_sheet("Performance")
sheet4.merge_cells('E1:I1')
sheet4.merge_cells('A1:D1')
performance_sheet_column_headers_1 = _performance_sheet_column_headers_1()
sheet4.append(performance_sheet_column_headers_1)
performance_sheet_column_headers_2 = _performance_sheet_column_headers_2()
sheet4.append(performance_sheet_column_headers_2)

#Alphabetize clusters
cluster_list = []
for cluster in performance_dictionary:
	cluster_list.append(cluster)
cluster_list.sort()

color_flag = 0
row_flag = 0

#Populate worksheet 
for cluster in cluster_list:
	cluster_flag = 1
	cluster_row_count = 0

	#Alphabetize host names
	host_name_list = []
	for host_name in performance_dictionary[cluster]:
		if host_name != None:
			host_name_list.append(host_name)
	host_name_list.sort()
	all_info = []

	#Alphabetize by Host Name
	for host_name in host_name_list:

		row_info_list = []
		for row_info in performance_dictionary[cluster][host_name]:
			row_info_list.append(row_info)
		
		#Populate spreadsheet
		all_info.append(cluster)
		all_info.append(host_name)
		for item in row_info_list:
			all_info.append(item)

		sheet4.append(all_info)
		all_info = []
	 
		#alternate colors for each cluster (white/blue)
		for row in sheet4.iter_rows(min_row=3, max_col=10):
			if row[0].value == cluster:
				for cell in row:
					if color_flag == 1:
						cell.fill = PatternFill(start_color="ffffff", fill_type="solid")
					if color_flag == 0:
						cell.fill = PatternFill(start_color="add8e6", fill_type="solid")
	if color_flag == 0:
		color_flag = 1
	elif color_flag == 1:
		color_flag = 0

	#get number of rows for each cluster
	for row in sheet4.iter_rows(min_row=3, max_col=1):
		for cell in row:
			if cell.value == cluster:
				cluster_row_count += 1	

	#apply borders as per cluster 
	for row in sheet4.iter_rows(min_row=3, max_col=10):
		if row[0].value == cluster and cluster_flag < cluster_row_count:	
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thin'))
				else:
					cell.border = Border(
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom 	= Side(style='thin'))

			cluster_flag += 1
				
		elif row[0].value == cluster and cluster_flag == cluster_row_count: 
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left 	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thick'))

				else:
					cell.border = Border(	
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom	= Side(style='thick'))

#Format header rows
for row in sheet4.iter_rows(min_row=1, max_row=2, max_col = 10):
	for cell in row:
		cell.fill 		= PatternFill(start_color="000080", fill_type="solid")
		cell.font 		= Font(bold=True, color="ffffff")
		cell.alignment 	= Alignment('center')
		cell.border 	= Border(
			left		= Side(style 	='thin',
								color 	="ffffff"),
			right 		= Side(style 	='thin',
								color 	="ffffff"),
			top			= Side(style 	='thin',
								color 	="ffffff"),
			bottom		= Side(style 	='thin',
								color 	="ffffff"))

#create proper widths for cells
for row in sheet4.iter_rows(max_row = 1):
	for cell in row:
		if cell.value:
			dimensions[cell.column] = max((dimensions.get(cell.column, 0), len(str(cell.value))+8))
for col, value in dimensions.items():
	sheet4.column_dimensions[col].width = value

#mark cpu % as red if over 80
for col in sheet4.iter_cols(min_col=5, max_col=5, min_row=3):
	for cell in col:
		if cell.value != "No Data Available":
			if cell.value >= 50:
				cell.font = Font(bold=True, color="ff0000")

#freeze top row
sheet4.freeze_panes='A3'

#add autofilter
column_letter 			= get_column_letter(len(performance_sheet_column_headers_2))
highest_row 			= sheet4.max_row
last_cell 				= column_letter + str(highest_row)
sheet4.auto_filter.ref 	= "A2:"  + last_cell



#################### VOLUMES SHEET #################### 

#Create sheet
sheet5 = wb.create_sheet("Volumes")
volumes_sheet_column_headers = _volumes_sheet_column_headers()
sheet5.append(volumes_sheet_column_headers)

#Alphabetize clusters
cluster_list = []
for cluster in volumes_dictionary:
	cluster_list.append(cluster)
cluster_list.sort()

color_flag = 0
row_flag = 0

#Populate worksheet 
for cluster in cluster_list:
	cluster_flag = 1
	cluster_row_count = 0

	#Alphabetize host names
	host_name_list = []
	for host_name in volumes_dictionary[cluster]:
		if host_name != None:
			host_name_list.append(host_name)
	host_name_list.sort()
	
	
	#Alphabetize by Host Name
	for host_name in host_name_list:

		row_info_list = []
		for row_info in volumes_dictionary[cluster][host_name]:
			row_info_list.append(row_info)
		row_info_list.sort(key=lambda x: float(x[3]), reverse=True)

		#Populate spreadsheet
		all_info = []
		for row in row_info_list:
			all_info.append(cluster)
			all_info.append(host_name)
			for item in row:
				all_info.append(item)
			sheet5.append(all_info)
			all_info = []
	 
		
		#alternate colors for each cluster (white/blue)
		for row in sheet5.iter_rows(min_row=2, max_col=6):
			if row[1].value == host_name:
				for cell in row:
					if color_flag == 1:
						cell.fill = PatternFill(start_color="ffffff", fill_type="solid")
					if color_flag == 0:
						cell.fill = PatternFill(start_color="add8e6", fill_type="solid")
		if color_flag == 0:
			color_flag = 1
		elif color_flag == 1:
			color_flag = 0

	#get number of rows for each cluster
	for row in sheet5.iter_rows(min_row=2, max_col=1):
		for cell in row:
			if cell.value == cluster:
				cluster_row_count += 1	

	#apply borders as per cluster 
	for row in sheet5.iter_rows(min_row=2, max_col=6):
		if row[0].value == cluster and cluster_flag < cluster_row_count:	
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thin'))
				else:
					cell.border = Border(
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom 	= Side(style='thin'))

			cluster_flag += 1
				
		elif row[0].value == cluster and cluster_flag == cluster_row_count: 
			for cell in row:
				if row.index(cell)<(len(row)-1):
					cell.border = Border(
						left 	= Side(style ='thin'),
						right 	= Side(style ='thin'),
						top 	= Side(style ='thin'),
						bottom 	= Side(style ='thick'))

				else:
					cell.border = Border(	
						left 	= Side(style='thin'),
						right 	= Side(style='thick'),
						top		= Side(style='thin'),
						bottom	= Side(style='thick'))

#add thin borders to all cells
for row in sheet5.iter_rows(min_row=1, max_row=1, max_col = 6):
	for cell in row:
		cell.fill 		= PatternFill(start_color="000080", fill_type="solid")
		cell.font 		= Font(bold=True, color="ffffff")
		cell.alignment 	= Alignment('center')
		cell.border 	= Border(
			left		= Side(style ='thin'),
			right 		= Side(style ='thin'),
			top			= Side(style ='thin'),
			bottom		= Side(style ='thin'))

#create proper widths for cells
for row in sheet5.iter_rows(max_row = 1):
	for cell in row:
		if cell.value:
			dimensions[cell.column] = max((dimensions.get(cell.column, 0), len(str(cell.value))+12))
for col, value in dimensions.items():
	sheet5.column_dimensions[col].width = value

#freeze top row
sheet3.freeze_panes='A2'

#add autofilter
column_letter 			= get_column_letter(len(capacity_trending_sheet_column_headers))
highest_row 			= sheet3.max_row
last_cell 				= column_letter + str(highest_row)
sheet3.auto_filter.ref 	= "A1:"  + last_cell



#################### SAVE EXCEL DOCUMENT #################### 

#Save new excel doc named 'NERD.xlsx'
wb.save('NERD.xlsx')

t1 = time.time()

total = (t1-t0)/float(60)
time_per_serial_num = (total*60)/(len(serial_numbers_list))

print "TOTAL TIME ELAPSED (MINUTES) = " + str(total)
print "AVG TIME PER SERIAL NUM (SECONDS) = " +  str(time_per_serial_num)